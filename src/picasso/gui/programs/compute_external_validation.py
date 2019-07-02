# -*- coding: utf-8 -*-
"""
Created on Thu May 24 14:31:48 2018

@author: sreimond
"""

from picasso.gui.programs import compute_grace
from picasso.utils.dates_and_time import date_functions
from picasso.utils.geometry import points
from picasso.gui import project_data_functions
import numpy as np
import os, sys
import openpyxl
import logging

data_path = None
density = None

def compute_external_validation(ui):
    # change button text
    logging.info('External validation computations started.')
    ui.pb_id004.setText('In progress...')
    ui.pb_id004.setStyleSheet("background-color: rgb(255, 250, 205); color: rgb(255,0,0);")
    # globals
    global data_path
    global density
    data_path = ui.le_id223.text()
    density = 1025.0
    if ui.rb_id073.isChecked():
        density = ui.dsb_id017.value()
    # dirs
    project_data_functions.create_project_directories(ui)
    project_data_functions.clean_up_external_validation_directories(ui)
    # analyze data
    _analyze_wgms(ui)
    # change back button text
    ui.pb_id004.setText('Compute External Validation')
    ui.pb_id004.setStyleSheet("background-color: rgb(239, 240, 241); color: rgb(0,0,0);")
    logging.info('External validation computations finished.')
    
def _analyze_wgms(ui):
    #-------------------------------------------------------------------------#
    fname_meta  = os.path.join(ui.le_id223.text(),'WGMS','Fog2017_10_04_MetaData.xlsx')
    fname_min   = os.path.join(ui.le_id223.text(),'WGMS','FoG2017_10_04_MinData.xlsx')
    fname_ref   = os.path.join(ui.le_id223.text(),'WGMS','REF-Glaciers_all.xlsx')
    global data_matrix
    #........................................................parse xlsx files #
    wb_meta = openpyxl.load_workbook(filename=fname_meta, read_only=True)
    mb_meta_header, mb_meta_data = sheet2dict(wb_meta['.._MB'])
    tc_meta_header, tc_meta_data = sheet2dict(wb_meta['.._TC'])
    wb_min = openpyxl.load_workbook(filename=fname_min, read_only=True)
    mb_min_header, mb_min_data = sheet2dict(wb_min['MinData_MB'])
    tc_min_header, tc_min_data = sheet2dict(wb_min['MinData_TC'])
    wb_ref = openpyxl.load_workbook(filename=fname_ref, read_only=True)
    ref_header, ref_data = sheet2dict(wb_ref['General Information'])    
    #........................................................fill data matrix #
    t_min = np.min((np.min(mb_meta_data['firstRY']),np.min(tc_meta_data['firstRY'])))
    t_max = np.max((np.max(mb_meta_data['lastSY']),np.max(tc_meta_data['lastSY'])))
    n_t = int(t_max)+1
    id_min = np.min((np.min(mb_meta_data['WGMS_ID']),np.min(tc_meta_data['WGMS_ID'])))
    id_max = np.max((np.max(mb_meta_data['WGMS_ID']),np.max(tc_meta_data['WGMS_ID'])))
    n_id = int(id_max)+1
    wgms_ids = range(0,n_id)
    data_types = ('ma','mw','ms','tc','vc')
    n_data_types = len(data_types)
    data_matrix = {}
    for d in data_types:
        data_matrix[d] = np.ones((n_id,n_t))*np.nan
    data_matrix['area'] = np.ones((n_id,n_t))*np.nan
    data_matrix['lon'] = np.ones((n_id,n_t))*np.nan
    data_matrix['lat'] = np.ones((n_id,n_t))*np.nan
    data_matrix['tr_mb'] = np.ones((n_id,n_t))*np.nan
    data_matrix['tr_tc'] = np.ones((n_id,n_t))*np.nan
    data_matrix['ma_from_tc'] = np.ones((n_id,n_t))*np.nan
    data_matrix['ma_from_vc'] = np.ones((n_id,n_t))*np.nan
    for i in range(len(mb_min_data['WGMS_ID'])):
        id_i = mb_min_data['WGMS_ID'][i]
        lon = mb_min_data['LONGITUDE'][i]
        lat = mb_min_data['LATITUDE'][i]
        ts = mb_min_data['SURVEY_YEAR'][i]
        tr = mb_min_data['REFERENCE_YEAR'][i]
        ma = mb_min_data['ANNUAL_BALANCE'][i]
        mw = mb_min_data['WINTER_BALANCE'][i]
        ms = mb_min_data['SUMMER_BALANCE'][i]
        data_matrix['lon'][id_i,ts] = lon
        data_matrix['lat'][id_i,ts] = lat
        data_matrix['ma'][id_i,ts] = ma
        data_matrix['mw'][id_i,ts] = mw
        data_matrix['ms'][id_i,ts] = ms
        data_matrix['tr_mb'][id_i,ts] = tr
    for i in range(len(tc_min_data['WGMS_ID'])):
        id_i = tc_min_data['WGMS_ID'][i]
        lon = mb_min_data['LONGITUDE'][i]
        lat = mb_min_data['LATITUDE'][i]
        area = tc_min_data['AREA_SURVEY_YEAR'][i]
        ts = tc_min_data['SURVEY_YEAR'][i]
        tr = tc_min_data['REFERENCE_YEAR'][i]
        tc = tc_min_data['THICKNESS_CHANGE'][i]
        tc = tc_min_data['TCcalc'][i]
        vc = tc_min_data['VOLUME_CHANGE'][i]
        data_matrix['lon'][id_i,ts] = lon
        data_matrix['lat'][id_i,ts] = lat
        data_matrix['area'][id_i,ts] = area
        data_matrix['tc'][id_i,ts] = tc
        data_matrix['vc'][id_i,ts] = vc
        data_matrix['tr_tc'][id_i,ts] = tr
    for id_i in range(id_min,id_max+1):
        for t in range(t_min,t_max+1):
            tc = data_matrix['tc'][id_i,t]
            tr = data_matrix['tr_tc'][id_i,t]
            if t==tr:
                tca = float(tc)
            else:
                tca = float(tc)/(t-tr)
            if np.isnan(tr):
                continue
            for ti in range(int(tr)+1,t+1):
                data_matrix['ma_from_tc'][id_i,ti] = tca*1e-3*density
            vc = data_matrix['vc'][id_i,t]
            tr = data_matrix['tr_tc'][id_i,t]
            area = data_matrix['area'][id_i,t]
            if t==tr:
                vca = float(vc)
            else:
                vca = float(vc)/(t-tr)
            if np.isnan(tr) or np.isnan(area) or area==0:
                continue
            for ti in range(int(tr)+1,t+1):
                data_matrix['ma_from_vc'][id_i,ti] = vca*density/(area*1e6)*1e3    
                data_matrix['area'][id_i,ti] = area
    #..............................................determine regional indices #
    polygon = compute_grace._region_settings(ui)
    glacier_region_code_mb = ui.cob_id004.currentText()
    glacier_region_code_tc = ui.cob_id005.currentText()
    wgms_ids_mb = np.array(ui.le_id007.text().split(),dtype=int)
    wgms_ids_ignore_mb = np.array(ui.le_id008.text().split(),dtype=int)
    wgms_ids_tc = np.array(ui.le_id009.text().split(),dtype=int)
    wgms_ids_ignore_tc = np.array(ui.le_id010.text().split(),dtype=int)
    use_only_refglaciers = False
    if ui.cb_id076.isChecked():
        use_only_refglaciers = True
    r_ix_glac = np.zeros(n_id,dtype=bool)
    r_ix_geod = np.zeros(n_id,dtype=bool)
    for i in range(len(mb_min_data['WGMS_ID'])):
        if use_only_refglaciers and (mb_min_data['WGMS_ID'][i] not in ref_data['WGMS_ID']):
            continue
        if ui.rb_id063.isChecked() and (mb_min_data['GLACIER_REGION_CODE'][i].lower() == glacier_region_code_mb.lower()):
            r_ix_glac[mb_min_data['WGMS_ID'][i]] = True
        if ui.rb_id064.isChecked():
            if polygon is None:
                r_ix_glac[mb_min_data['WGMS_ID'][i]] = True
            else:
                point = points.Point2D(mb_min_data['LONGITUDE'][i],mb_min_data['LATITUDE'][i])
                if polygon.determine_point_location(point)>0:
                    r_ix_glac[mb_min_data['WGMS_ID'][i]] = True
        if ui.rb_id065.isChecked():
            if use_only_refglaciers and (mb_min_data['WGMS_ID'][i] not in ref_data['WGMS_ID']):
                continue
            if mb_min_data['WGMS_ID'][i] in wgms_ids_mb:
                r_ix_glac[mb_min_data['WGMS_ID'][i]] = True
    for i in range(len(tc_min_data['WGMS_ID'])):
        if use_only_refglaciers and (tc_min_data['WGMS_ID'][i] not in ref_data['WGMS_ID']):
            continue
        if ui.rb_id066.isChecked() and (tc_min_data['GLACIER_REGION_CODE'][i].lower() == glacier_region_code_tc.lower()):
            r_ix_geod[tc_min_data['WGMS_ID'][i]] = True 
        if ui.rb_id067.isChecked():
            if polygon is None:
                r_ix_geod[tc_min_data['WGMS_ID'][i]] = True
            else:
                point = points.Point2D(tc_min_data['LONGITUDE'][i],tc_min_data['LATITUDE'][i])
                if polygon.determine_point_location(point)>0:
                    r_ix_geod[tc_min_data['WGMS_ID'][i]] = True
        if ui.rb_id068.isChecked():
            if use_only_refglaciers and (tc_min_data['WGMS_ID'][i] not in ref_data['WGMS_ID']):
                continue
            if tc_min_data['WGMS_ID'][i] in wgms_ids_tc:
                r_ix_geod[mb_min_data['WGMS_ID'][i]] = True
    #.....................................determine indices to ultimately use #
    r_ix_glac_avail = np.copy(r_ix_glac)
    r_ix_geod_avail = np.copy(r_ix_geod)
    for i in range(n_id):
        if r_ix_glac[i] in wgms_ids_ignore_mb:
            r_ix_glac[i] = False
        if r_ix_geod[i] in wgms_ids_ignore_tc:
            r_ix_geod[i] = False
    #.............................................calculate regional balances #
    epoch_glac = (int(ui.de_id007.date().toString("yyyy")),int(ui.de_id008.date().toString("yyyy")))
    epoch_geod = (int(ui.de_id005.date().toString("yyyy")),int(ui.de_id006.date().toString("yyyy")))
    mass_balance_matrix = {}
    for d in data_types:
        mass_balance_matrix[d] = np.ones(n_t)*np.nan
        did = d+'_ids_avail'
        mass_balance_matrix[did] = np.ones((n_id,n_t))*np.nan
        did = d+'_ids_used'
        mass_balance_matrix[did] = np.ones((n_id,n_t))*np.nan
    mass_balance_matrix['ma_from_tc'] = np.ones(n_t)*np.nan
    mass_balance_matrix['ma_from_vc'] = np.ones(n_t)*np.nan
    mass_balance_matrix['nobs_ma'] = np.ones(n_t)*np.nan
    mass_balance_matrix['nobs_mw'] = np.ones(n_t)*np.nan
    mass_balance_matrix['nobs_ms'] = np.ones(n_t)*np.nan
    mass_balance_matrix['nobs_tc'] = np.ones(n_t)*np.nan
    mass_balance_matrix['nobs_vc'] = np.ones(n_t)*np.nan
    mass_balance_matrix['area'] = np.ones(n_t)*np.nan
    for t in range(epoch_glac[0],epoch_glac[1]+1):
        if t>t_max:
            continue
        for key in ('ma','mw','ms'):
            yi = data_matrix[key][r_ix_glac,t]
            nobs = np.count_nonzero(~np.isnan(yi))
            mass_balance_matrix['nobs_'+key][t] = nobs
            if nobs > 0:
                mass_balance_matrix[key][t] = np.nanmean(yi)
            tmp = np.array(~np.isnan(yi),dtype=np.float_)
            tmp[tmp==0] = np.nan
            mass_balance_matrix[key+'_ids_used'][r_ix_glac,t] = tmp
    for t in range(epoch_geod[0],epoch_geod[1]+1):
        if t>t_max:
            continue
        for key in ('tc','vc'):
            yi = data_matrix['ma_from_'+key][r_ix_geod,t]
            ai = data_matrix['area'][r_ix_geod,t]
            nobs = np.count_nonzero(~np.isnan(yi))
            mass_balance_matrix['nobs_'+key][t] = nobs
            if nobs > 0:
                mass_balance_matrix['ma_from_'+key][t] = np.nansum(yi*ai/np.nansum(ai))
                # mass_balance_matrix['ma_from_'+key][t] = np.nanmean(yi)
            tmp = np.array(~np.isnan(yi),dtype=np.float_)
            tmp[tmp==0] = np.nan
            mass_balance_matrix[key+'_ids_used'][r_ix_geod,t] = tmp
        mass_balance_matrix['area'][t] = np.nansum(data_matrix['area'][r_ix_geod,t])
    for t in range(t_min,t_max+1): 
        for key in ('ma','mw','ms'):
            yi = data_matrix[key][r_ix_glac_avail,t]
            tmp = np.array(~np.isnan(yi),dtype=np.float_)
            tmp[tmp==0] = np.nan
            mass_balance_matrix[key+'_ids_avail'][r_ix_glac_avail,t] = tmp
        for key in ('tc','vc'):
            yi = data_matrix['ma_from_'+key][r_ix_geod_avail,t]
            tmp = np.array(~np.isnan(yi),dtype=np.float_)
            tmp[tmp==0] = np.nan
            mass_balance_matrix[key+'_ids_avail'][r_ix_geod_avail,t] = tmp
    #............................................calibrate glaciological data #
    tmp = np.copy(mass_balance_matrix['ma'])
    tmp -= np.nanmean(tmp)
    tmp += np.nanmean(mass_balance_matrix['ma_from_tc'])
    mass_balance_matrix['ma_calibrated'] = tmp
    #...............................................................save data #
    save_file(ui,t_max,mass_balance_matrix)
    save_file_cumsum(ui,t_max,mass_balance_matrix)

def sheet2dict(sheet):
	dh = {}
	data = {}
	ix = -1
	for row in sheet.iter_rows(min_row=1,max_row=1):
		for cell in row:
			ix += 1
			dh[ix] = cell.value
			data[cell.value] = []
	for row in sheet.iter_rows(min_row=2):
		ix = -1
		for cell in row:
			ix += 1
			data[dh[ix]].append(cell.value)	
	return dh, data
              
def save_file(ui,t_max,mb_matrix):
    project_name = str(ui.le_id001.text())
    project_dir = str(ui.le_id002.text())
    external_validation_path = os.path.join(project_dir,project_name,'validation','external')
    fname = os.path.join(external_validation_path,'mass_balances','WGMS_mass_balance.txt')
    epoch_glac = (int(ui.de_id007.date().toString("yyyy")),int(ui.de_id008.date().toString("yyyy")))
    epoch_geod = (int(ui.de_id005.date().toString("yyyy")),int(ui.de_id006.date().toString("yyyy")))
    tlims = [np.min((epoch_glac[0],epoch_geod[0])),np.max((epoch_glac[1],epoch_geod[1]))]
    if tlims[1]>t_max:
        tlims[1] = t_max
    t = range(tlims[0],tlims[1]+1)
    keys = ('mw',
            'ms',
            'ma',
            'ma_from_tc',
            'ma_from_vc',
            'ma_calibrated',
            'nobs_mw',
            'nobs_ms',
            'nobs_ma',
            'nobs_tc',
            'nobs_vc',
            'area')
    fw = open(fname,'w')
    fw.write('%14s' % 't')
    for key in keys:
        fw.write('%14s' % key)
    fw.write('\n')
    for ti in t:
        fw.write('%14d' % ti)
        for key in keys:
            if 'nobs' in key:
                fw.write('%14.0f' % mb_matrix[key][ti])
            else:
                fw.write('%14f' % mb_matrix[key][ti])
        fw.write('\n')
    fw.close()
#    
def save_file_cumsum(ui,t_max,mb_matrix):     
    project_name = str(ui.le_id001.text())
    project_dir = str(ui.le_id002.text())
    external_validation_path = os.path.join(project_dir,project_name,'validation','external')
    epoch_glac = (int(ui.de_id007.date().toString("yyyy")),int(ui.de_id008.date().toString("yyyy")))
    epoch_geod = (int(ui.de_id005.date().toString("yyyy")),int(ui.de_id006.date().toString("yyyy")))
    tlims = [np.min((epoch_glac[0],epoch_geod[0])),np.max((epoch_glac[1],epoch_geod[1]))]
    if tlims[1]>t_max:
        tlims[1] = t_max
    t = range(tlims[0],tlims[1]+1)
    # area
    area = np.copy(mb_matrix['area'])
    if ui.rb_id069.isChecked():
        area = area*0 + ui.dsb_id001.value()
    elif ui.rb_id071.isChecked():
        area = area*0 + ui.dsb_id016.value()
    mjd = []
    kg = []
    kg_cal = []
    for ti in t:
        t = date_functions.ymd2mjd( ti, 10, 1 )
        yi =  mb_matrix['ma'][ti] * (area[ti]*1e6) * (density/1e3)                 # 1 m w.e. = 1000 kg m^-2 rho^-1
        yi_cal =  mb_matrix['ma_calibrated'][ti] * (area[ti]*1e6) * (density/1e3)
        mjd.append(t)
        kg.append(yi)
        kg_cal.append(yi_cal)
    kg = np.nancumsum(kg)
    kg_cal = np.nancumsum(kg_cal)
    ix = kg==0
    kg[ix] = np.nan
    ix = kg_cal==0
    kg_cal[ix] = np.nan
    data = {'wgms':kg,'wgms_calibrated':kg_cal}
    for key in ('wgms','wgms_calibrated'):
        time_series_file = os.path.join(external_validation_path,'time_series','%s.txt' % key)
        with open(time_series_file,'w') as f:
            f.write('%25s %25s %25s\n' % ('time (mjd)','mass (kg)','sigma (kg)'))
        for ix in range(len(kg)):
            mjdi = mjd[ix]
            with open(time_series_file,'a+') as f:
                f.write('%+25.16e %+25.16e %+25.16e\n' % (mjdi,data[key][ix],0))
                




                